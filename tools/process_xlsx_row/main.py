"""process_xlsx_row: highlight xlsx rows that match a boolean expression of rules.

Usage:
  python -m tools process-xlsx-row data.xlsx --bgColor yellow --rules "r1|r2" \
      --rules-file rules.json
  python -m tools process-xlsx-row data.xlsx --bgColor "#FFFF00" --rules "r1&!r3" \
      --rules-script rules.py
"""
import argparse
import copy
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import PatternFill

from tools.process_xlsx_row.rules import (
    RuleRegistry,
    SafeExpressionEvaluator,
    load_rules_from_json,
    load_rules_from_script,
)


# Color name -> 8-char ARGB hex (openpyxl/Excel format: AARRGGBB)
NAMED_COLORS: Dict[str, str] = {
    'black': 'FF000000',
    'white': 'FFFFFFFF',
    'red': 'FFFF0000',
    'green': 'FF00FF00',
    'blue': 'FF0000FF',
    'yellow': 'FFFFFF00',
    'cyan': 'FF00FFFF',
    'magenta': 'FFFF00FF',
    'orange': 'FFFFA500',
    'purple': 'FF800080',
    'pink': 'FFFFC0CB',
    'gray': 'FF808080',
    'grey': 'FF808080',
    'lightgray': 'FFD3D3D3',
    'lightgrey': 'FFD3D3D3',
    'darkgray': 'FFA9A9A9',
    'darkgrey': 'FFA9A9A9',
    'brown': 'FFA52A2A',
    'navy': 'FF000080',
    'teal': 'FF008080',
    'olive': 'FF808000',
    'lime': 'FF00FF00',
    'maroon': 'FF800000',
    'silver': 'FFC0C0C0',
    'gold': 'FFFFD700',
}


def parse_color(value: str) -> str:
    """Parse a color to 8-char ARGB hex. Accepts named colors, #RRGGBB, AARRGGBB."""
    s = value.strip()
    if not s:
        raise ValueError("color value cannot be empty")
    if s.lower() in NAMED_COLORS:
        return NAMED_COLORS[s.lower()]
    s = s.lstrip('#')
    if re.fullmatch(r'[0-9A-Fa-f]{6}', s):
        return 'FF' + s.upper()
    if re.fullmatch(r'[0-9A-Fa-f]{8}', s):
        return s.upper()
    raise ValueError(
        f"invalid color {value!r}; expected a named color (e.g. 'yellow', "
        f"'red'), #RRGGBB, or AARRGGBB hex"
    )


def _load_rules(
    rules_file: Optional[Path],
    rules_script: Optional[Path],
) -> RuleRegistry:
    registry = RuleRegistry()
    if rules_file:
        load_rules_from_json(rules_file, registry)
    if rules_script:
        load_rules_from_script(rules_script, registry)
    if not registry.rules:
        raise ValueError(
            "no rules loaded; provide --rules-file and/or --rules-script"
        )
    return registry


def _resolve_output_path(
    input_path: Path,
    output: Optional[Path],
    in_place: bool,
) -> Path:
    if in_place and output:
        raise ValueError("use either --in-place or --output, not both")
    if in_place:
        return input_path
    if output:
        return Path(output).expanduser()
    return input_path.with_name(f"{input_path.stem}_colored{input_path.suffix}")


def _color_entire_row(
    ws,
    row_idx: int,
    max_col: int,
    fill: PatternFill,
) -> None:
    """Apply `fill` to every cell in `row_idx`, columns 1..max_col.

    The whole row is colored (not just the columns referenced by rules), so a
    matched row is visually unambiguous regardless of which rule triggered it.
    `max_col` should be the sheet's `ws.max_column` captured once up front.
    """
    for c in range(1, max_col + 1):
        ws.cell(row=row_idx, column=c).fill = fill


def _export_matched_rows(
    src_path: Path,
    sheet_name: str,
    matched_indices: List[int],
    header_rows: int,
    fill: PatternFill,
    output_path: Path,
) -> None:
    """Write a new xlsx containing `header_rows` header rows + matched data rows.

    Matched data rows are written with the highlight `fill` applied to every
    cell, mirroring the visual cue of the main colored workbook. Cell values,
    basic style (font/alignment/number_format), and column widths are copied
    from the source. The source workbook is re-read from disk (not the in-memory
    `wb` that already has colors applied), so the only color in the export is
    the explicit highlight on matched rows.
    """
    src_wb = openpyxl.load_workbook(str(src_path), data_only=True)
    if sheet_name in src_wb.sheetnames:
        src_ws = src_wb[sheet_name]
    else:
        src_ws = src_wb.active
    max_col = src_ws.max_column

    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = src_ws.title

    def _copy_cell(src_cell, dst_cell):
        dst_cell.value = src_cell.value
        if src_cell.has_style:
            dst_cell.font = copy.copy(src_cell.font)
            dst_cell.alignment = copy.copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format

    for r in range(1, header_rows + 1):
        for c in range(1, max_col + 1):
            _copy_cell(src_ws.cell(row=r, column=c), new_ws.cell(row=r, column=c))

    dst_row = header_rows + 1
    for src_row in matched_indices:
        for c in range(1, max_col + 1):
            _copy_cell(
                src_ws.cell(row=src_row, column=c),
                new_ws.cell(row=dst_row, column=c),
            )
        for c in range(1, max_col + 1):
            new_ws.cell(row=dst_row, column=c).fill = fill
        dst_row += 1

    for key, dim in src_ws.column_dimensions.items():
        if dim.width is not None:
            new_ws.column_dimensions[key].width = dim.width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    new_wb.save(str(output_path))


def run(args: argparse.Namespace) -> int:
    input_path = Path(args.xlsx).expanduser().resolve()
    if not input_path.exists():
        print(f"error: file not found: {input_path}", file=sys.stderr)
        return 1
    if input_path.suffix.lower() != '.xlsx':
        print(
            f"warning: input is not .xlsx ({input_path.suffix}); "
            f"openpyxl may not handle it correctly",
            file=sys.stderr,
        )

    try:
        color = parse_color(args.bgColor)
    except ValueError as e:
        print(f"error: {e}", file=sys.stderr)
        return 1

    try:
        registry = _load_rules(args.rules_file, args.rules_script)
    except (ValueError, RuntimeError, FileNotFoundError) as e:
        print(f"error: {e}", file=sys.stderr)
        return 1

    try:
        evaluator = SafeExpressionEvaluator(args.rules, registry)
    except ValueError as e:
        print(f"error: {e}", file=sys.stderr)
        return 1

    try:
        wb = openpyxl.load_workbook(str(input_path), data_only=True)
    except Exception as e:
        print(f"error: failed to open xlsx: {e}", file=sys.stderr)
        return 1

    if args.sheet:
        if args.sheet not in wb.sheetnames:
            print(
                f"error: sheet {args.sheet!r} not found; "
                f"available: {wb.sheetnames}",
                file=sys.stderr,
            )
            return 1
        ws = wb[args.sheet]
    else:
        ws = wb.active

    if ws.max_row < 1:
        print("error: worksheet has no rows", file=sys.stderr)
        return 1

    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    max_col = max(ws.max_column, 1)
    start_row = 2 if args.header else 1

    total = 0
    matched = 0
    rule_hits: Dict[str, int] = {rid: 0 for rid in registry.rules}
    sample_matches: List[int] = []
    matched_indices: List[int] = []
    progress_step = 1000

    for row_idx in range(start_row, ws.max_row + 1):
        row_values = [
            ws.cell(row=row_idx, column=c).value
            for c in range(1, max_col + 1)
        ]
        total += 1
        try:
            ok, results = evaluator.evaluate(row_values)
        except RuntimeError as e:
            print(f"error at row {row_idx}: {e}", file=sys.stderr)
            return 1
        for rid, hit in results.items():
            if hit:
                rule_hits[rid] += 1
        if ok:
            matched += 1
            matched_indices.append(row_idx)
            _color_entire_row(ws, row_idx, max_col, fill)
            if len(sample_matches) < 5:
                sample_matches.append(row_idx)
        if total % progress_step == 0:
            print(
                f"  ... processed {total} rows, {matched} matched",
                file=sys.stderr,
            )

    if args.dry_run:
        out_path = input_path
        save_note = "(dry run, file not written)"
    else:
        try:
            out_path = _resolve_output_path(
                input_path, args.output, args.in_place
            )
            if out_path != input_path:
                out_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(out_path))
            save_note = ""
        except (ValueError, OSError) as e:
            print(f"error: {e}", file=sys.stderr)
            return 1

    export_path: Optional[Path] = None
    export_note = ""
    if args.export_matches:
        if args.dry_run:
            export_path = Path(args.export_matches).expanduser()
            export_note = "(dry run, file not written)"
        elif not matched_indices:
            print(
                "warning: --export-matches given but 0 rows matched; "
                "no export file written",
                file=sys.stderr,
            )
        else:
            try:
                export_path = Path(args.export_matches).expanduser().resolve()
                if args.output:
                    out_resolved = Path(args.output).expanduser().resolve()
                    if export_path == out_resolved:
                        print(
                            "error: --export-matches and --output point to "
                            "the same path",
                            file=sys.stderr,
                        )
                        return 1
                _export_matched_rows(
                    input_path,
                    ws.title,
                    matched_indices,
                    1 if args.header else 0,
                    fill,
                    export_path,
                )
            except (OSError, ValueError) as e:
                print(f"error: failed to export matches: {e}", file=sys.stderr)
                return 1

    print(f"input:        {input_path}")
    print(f"sheet:        {ws.title}")
    rule_ids = sorted(registry.rules.keys())
    print(
        f"rules:        {args.rules}  "
        f"({len(rule_ids)} loaded: {', '.join(rule_ids)})"
    )
    rule_cols = ", ".join(
        f"{rid}={list(registry.rules[rid].columns)}"
        for rid in rule_ids
    )
    print(f"rule columns: {rule_cols}")
    print(f"color:        {args.bgColor} -> {color}")
    if rule_hits:
        hits_str = ", ".join(f"{rid}={n}" for rid, n in rule_hits.items())
        print(f"per-rule hits:{hits_str}")
    print(f"rows scanned: {total}")
    print(f"rows matched: {matched}")
    if sample_matches:
        suffix = ", ..." if matched > len(sample_matches) else ""
        print(
            f"sample rows:  "
            f"{', '.join(str(r) for r in sample_matches)}{suffix}"
        )
    print(f"output:       {out_path}  {save_note}".rstrip())
    if export_path:
        print(
            f"export:       {export_path}  "
            f"({len(matched_indices)} rows){(' ' + export_note) if export_note else ''}"
        )
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog='process-xlsx-row',
        description=(
            'Set row background color in an xlsx file based on a boolean '
            'expression of cell-matching rules.'
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "expression syntax for --rules:\n"
            "  r1|r2             OR       (or: r1 or r2)\n"
            "  r1&r2             AND      (or: r1 and r2)\n"
            "  !r1               NOT r1   (or: not r1)\n"
            "  (r1|r2)&!r3       parentheses, precedence\n"
            "\n"
            "rule file (--rules-file, JSON):\n"
            '  {"r1": {"column": 1, "match": "contains:Active"}}\n'
            '  {"r2": {"column": 2, "match": {"type": "regex", "value": "^foo"}}}\n'
            '  {"r3": {"column": 3, "match": {"type": "py", "function": "my_check"}}}\n'
            "\n"
            "rule script (--rules-script, Python):\n"
            "  def my_check(v): return str(v).startswith('A')\n"
            "  RULES = {\n"
            "      'r1': {'column': 1, 'match': 'equals:Active'},\n"
            "      'r3': {'column': 3, 'match': my_check},\n"
            "  }\n"
            "\n"
            "built-in matchers: equals, contains, startswith, endswith, regex\n"
            "\n"
            "color: name (yellow, red, ...) or #RRGGBB / AARRGGBB hex"
        ),
    )
    parser.add_argument('xlsx', help='Path to input .xlsx file')
    parser.add_argument(
        '--bgColor', required=True,
        help='Background color (e.g. "yellow", "#FFFF00", "FFFFFF00")',
    )
    parser.add_argument(
        '--rules', required=True,
        help='Boolean expression over rule IDs (e.g. "r1|r2", "r1&!r3")',
    )
    parser.add_argument(
        '--rules-file', type=Path, default=None,
        help='JSON file mapping rule_id -> {column, match}',
    )
    parser.add_argument(
        '--rules-script', type=Path, default=None,
        help='Python script defining RULES (dict or list) and optional '
             'custom match functions for "py" matchers',
    )
    parser.add_argument(
        '--sheet', default=None,
        help='Worksheet name (default: active sheet)',
    )
    parser.add_argument(
        '--header', action='store_true',
        help='Treat the first row as a header (skip from matching)',
    )
    parser.add_argument(
        '--output', type=Path, default=None,
        help='Output xlsx path (default: <input>_colored.xlsx next to input)',
    )
    parser.add_argument(
        '--in-place', action='store_true',
        help='Overwrite the input file (default: write to <input>_colored.xlsx)',
    )
    parser.add_argument(
        '--dry-run', action='store_true',
        help='Show what would be done without writing any file',
    )
    parser.add_argument(
        '--export-matches', type=Path, default=None, metavar='PATH',
        help='Also write a separate xlsx containing the matched rows (with '
             'the highlight color preserved). The header row is included '
             'only when --header is set.',
    )
    return parser


def main(argv: Optional[List[str]] = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    return run(args)


if __name__ == '__main__':
    sys.exit(main())
