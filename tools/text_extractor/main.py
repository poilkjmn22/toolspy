#!/usr/bin/env python3
import argparse
import shlex
import sys
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Tuple

import openpyxl
import pdfplumber
import pypdfium2 as pdfium
import xlrd

from tools.ocr_engine import (
    OCREngine,
    check_engine,
    get_engine,
)


SUPPORTED_EXTS = {'.pdf', '.xls', '.xlsx'}
FILE_HEADER_TEMPLATE = "# Extracted from {name} by text-extractor\n"
OCR_HEADER_TEMPLATE = (
    "# Extracted from {name} by text-extractor "
    "(OCR via {engine}, lang={lang}, dpi={dpi})\n"
)


def check_tesseract() -> Tuple[bool, str]:
    """Backwards-compat wrapper for `check_engine('tesseract')`."""
    return check_engine('tesseract')


def check_paddleocr() -> Tuple[bool, str]:
    """Backwards-compat wrapper for `check_engine('paddleocr')`."""
    return check_engine('paddleocr')


def _row_is_empty(row) -> bool:
    for c in row:
        if c is None:
            continue
        if isinstance(c, str) and c.strip() == "":
            continue
        return False
    return True


def _pdfplumber_extract(path: Path, warn: bool = True) -> str:
    parts = [FILE_HEADER_TEMPLATE.format(name=path.name)]
    total_chars = 0
    page_count = 0
    with pdfplumber.open(str(path)) as pdf:
        page_count = len(pdf.pages)
        for i, page in enumerate(pdf.pages, 1):
            parts.append(f"\n=== Page {i} ===\n\n")
            text = page.extract_text() or ""
            parts.append(text)
            total_chars += len(text)
    if warn and total_chars == 0 and page_count > 0:
        print(
            f"⚠ {path}: no text layer ({page_count} pages, 0 chars); "
            f"use --ocr to extract via OCR",
            file=sys.stderr,
        )
    return "".join(parts)


def _ocr_extract(path: Path, lang: str, dpi: int,
                 rotation: Optional[int] = None,
                 engine: Optional[OCREngine] = None) -> str:
    """OCR a PDF via the supplied (or default) engine.

    If `rotation` is None (default), try both orientations and pick the
    better result. If set to 0/90/180/270 (degrees, CW = positive), force
    that rotation. PaddleOCR (with use_angle_cls=True) handles rotation
    internally — the rotation arg is ignored in that case.

    If `engine` is None, defaults to TesseractEngine('chi_sim+eng').
    """
    if engine is None:
        engine = get_engine('tesseract', lang=lang)
        engine.init()
    parts = [OCR_HEADER_TEMPLATE.format(
        name=path.name, lang=lang, dpi=dpi, engine=engine.name)]
    pdf = pdfium.PdfDocument(str(path))
    try:
        scale = dpi / 72.0
        for i, page in enumerate(pdf, 1):
            parts.append(f"\n=== Page {i} ===\n\n")
            pil = page.render(scale=scale).to_pil()
            try:
                if rotation is not None:
                    # Manual rotation (CW = positive in user-facing convention).
                    # Ignored for PaddleOCR (it rotates via angle_cls internally).
                    if rotation != 0:
                        pil = pil.rotate(-rotation, expand=True)
                    text = engine.ocr(pil)
                    if rotation != 0:
                        parts.append(f"# OCR rotated {rotation}° CW (forced)\n")
                    parts.append(text)
                else:
                    # Auto: try original + 90° CW, pick better.
                    # For PaddleOCR (use_angle_cls=True) the second pass is
                    # redundant — but cheap, and keeps the behavior uniform.
                    text_default = engine.ocr(pil)
                    pil_rot = pil.rotate(-90, expand=True)
                    try:
                        text_rotated = engine.ocr(pil_rot)
                    finally:
                        pil_rot.close()
                    text, used_rot = _pick_better_text(text_default, text_rotated)
                    if used_rot:
                        parts.append("# OCR auto-rotated 90° CW (text was vertical in original)\n")
                    parts.append(text)
            finally:
                pil.close()
    finally:
        pdf.close()
    return "".join(parts)


def _detect_rotation(pil: 'PIL.Image.Image', engine: Optional[OCREngine] = None) -> tuple:
    """Detect page rotation via the engine's own OSD. Returns (confidence, degrees)."""
    if engine is None:
        engine = get_engine('tesseract')
        engine.init()
    return engine.detect_rotation(pil)


def _pick_better_text(text_a: str, text_b: str) -> tuple:
    """Pick the OCR output with more Chinese characters (better text recognition).
    Returns (text, used_b).
    """
    def chinese_count(s):
        return sum(1 for c in s if '\u4e00' <= c <= '\u9fff')
    ca = chinese_count(text_a)
    cb = chinese_count(text_b)
    # Prefer rotated only if it has 30%+ more Chinese (to avoid noise)
    if cb > ca * 1.3 and cb > 50:
        return (text_b, True)
    return (text_a, False)


def _xlsx_extract(path: Path) -> str:
    parts = [FILE_HEADER_TEMPLATE.format(name=path.name)]
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"\n=== Sheet: {sheet_name} ===\n\n")
            for row in ws.iter_rows(values_only=True):
                if _row_is_empty(row):
                    continue
                parts.append("\t".join("" if c is None else str(c) for c in row))
                parts.append("\n")
    finally:
        wb.close()
    return "".join(parts)


def _xls_extract(path: Path) -> str:
    parts = [FILE_HEADER_TEMPLATE.format(name=path.name)]
    book = xlrd.open_workbook(str(path))
    for sheet in book.sheets():
        parts.append(f"\n=== Sheet: {sheet.name} ===\n\n")
        for row_idx in range(sheet.nrows):
            row = sheet.row_values(row_idx)
            if _row_is_empty(row):
                continue
            parts.append("\t".join("" if c is None else str(c) for c in row))
            parts.append("\n")
    return "".join(parts)


def extract_text(
    path,
    ocr: bool = False,
    lang: str = 'chi_sim+eng',
    dpi: int = 300,
    warn: bool = True,
    rotation: Optional[int] = None,
    engine: Optional[OCREngine] = None,
) -> str:
    """Extract text from a single file. Convenience function for use by other tools.

    Args:
        path: file path (PDF/XLS/XLSX)
        ocr: if True, use OCR for PDFs (slower; works on scanned/image-only PDFs)
        lang: OCR language packs (e.g. 'chi_sim+eng', 'eng')
        dpi: render DPI for OCR (150-400)
        warn: if True, print a warning to stderr when a PDF has no text layer and ocr=False
        rotation: if set, force PDF page rotation in degrees (0/90/180/270, CW=positive).
                   Default None = auto-detect (try both orientations, pick the better one).
        engine: optional pre-built OCREngine instance. If None and ocr=True,
                a TesseractEngine is created on the fly. Pass an existing engine
                to avoid per-file initialization cost (model load for PaddleOCR).

    Returns:
        Extracted text with file header and per-page/per-sheet markers.
    """
    path = Path(path)
    ext = path.suffix.lower()
    if ext == '.pdf':
        if ocr:
            return _ocr_extract(path, lang, dpi, rotation=rotation, engine=engine)
        return _pdfplumber_extract(path, warn=warn)
    if ext == '.xlsx':
        return _xlsx_extract(path)
    if ext == '.xls':
        return _xls_extract(path)
    raise ValueError(f"不支持的扩展名: {ext}")


class TextExtractor:
    def __init__(
        self,
        sources: List[str],
        output: Optional[str] = None,
        include_subfolders: bool = True,
        combine: bool = False,
        overwrite: bool = False,
        ocr: bool = False,
        lang: str = 'chi_sim+eng',
        dpi: int = 300,
        rotation: Optional[int] = None,
        engine: str = 'tesseract',
        use_gpu: bool = False,
    ):
        self.sources = [Path(s).expanduser() for s in sources]
        self.output = Path(output).expanduser() if output else None
        self.include_subfolders = include_subfolders
        self.combine = combine
        self.overwrite = overwrite
        self.ocr = ocr
        self.lang = lang
        self.dpi = dpi
        self.rotation = rotation
        self.engine_name = engine
        self.use_gpu = use_gpu
        self._engine: Optional[OCREngine] = None
        self.stats = {
            'total': 0,
            'processed': 0,
            'skipped': 0,
            'failed': 0,
            'start_time': datetime.now(),
        }
        self.failures: List[Tuple[Path, str]] = []



    def _default_output_dir(self) -> Path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        base = self.sources[0].parent if self.sources else Path.cwd()
        return base / f"text_extracted_{ts}"

    def validate(self) -> Tuple[bool, str]:
        if not self.sources:
            return False, "至少需要一个源文件或文件夹"

        for s in self.sources:
            if not s.exists():
                return False, f"源路径不存在: {s}"

        if self.combine:
            if self.output is None:
                return False, "--combine 模式必须通过 -o/--output 指定输出 .txt 文件"
        else:
            if self.output is None:
                self.output = self._default_output_dir()
                print(f"输出目录: {self.output}")

        if not (150 <= self.dpi <= 400):
            return False, f"--dpi 必须在 150-400 之间，当前: {self.dpi}"

        return True, "验证通过"

    def _prepare_output(self) -> None:
        if self.combine:
            self.output.parent.mkdir(parents=True, exist_ok=True)
        else:
            self.output.mkdir(parents=True, exist_ok=True)

    def discover(self) -> List[Tuple[Path, Path, Optional[Path]]]:
        items: List[Tuple[Path, Path, Optional[Path]]] = []
        for src in self.sources:
            if src.is_file():
                if src.suffix.lower() in SUPPORTED_EXTS:
                    items.append((src, self._output_path_for(src), None))
            elif src.is_dir():
                pattern = "**/*" if self.include_subfolders else "*"
                for p in sorted(src.glob(pattern)):
                    if p.is_file() and p.suffix.lower() in SUPPORTED_EXTS:
                        items.append((p, self._output_path_for(p, base_folder=src), src))
        return items

    def _output_path_for(self, input_path: Path, base_folder: Optional[Path] = None) -> Path:
        if self.combine:
            return self.output
        base = base_folder if base_folder is not None else input_path.parent
        rel = input_path.relative_to(base)
        return self.output / rel.with_suffix('.txt')

    def _extract(self, path: Path) -> str:
        return extract_text(path, ocr=self.ocr, lang=self.lang, dpi=self.dpi,
                            warn=True, rotation=self.rotation, engine=self._get_engine())

    def _get_engine(self) -> Optional[OCREngine]:
        """Lazy-build the OCR engine on first use (model load is expensive)."""
        if not self.ocr:
            return None
        if self._engine is None:
            try:
                self._engine = get_engine(self.engine_name, lang=self.lang, use_gpu=self.use_gpu)
                self._engine.init()
            except Exception as e:
                print(f'Warning: failed to init engine {self.engine_name}: {e}', file=sys.stderr)
                # Fall back to Tesseract
                self._engine = get_engine('tesseract', lang=self.lang)
                self._engine.init()
        return self._engine

    def _write_output(self, out_path: Path, content: str) -> None:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        with open(out_path, 'w', encoding='utf-8', errors='replace') as f:
            f.write(content)

    def convert(self) -> Tuple[bool, str]:
        valid, msg = self.validate()
        if not valid:
            return False, msg

        if self.ocr:
            ok, tmsg = check_engine(self.engine_name)
            if not ok:
                return False, tmsg
            print(f"OCR 引擎: {tmsg}")

        items = self.discover()
        self.stats['total'] = len(items)
        if not items:
            return False, f"未找到可处理的文件 (支持: {', '.join(sorted(SUPPORTED_EXTS))})"

        self._prepare_output()
        print(f"找到 {self.stats['total']} 个待处理文件")

        combined_handle = None
        try:
            if self.combine:
                if self.output.exists() and not self.overwrite:
                    return False, f"输出文件已存在: {self.output} (使用 --overwrite 覆盖)"
                combined_handle = open(self.output, 'w', encoding='utf-8', errors='replace')

            for idx, (src, out, _base) in enumerate(items, 1):
                print(f"处理文件 {idx}/{self.stats['total']}: {src.name}")

                if not self.combine and out.exists() and not self.overwrite:
                    print(f"  跳过: 输出已存在 {out} (使用 --overwrite 覆盖)")
                    self.stats['skipped'] += 1
                    continue

                try:
                    content = self._extract(src)
                    if self.combine:
                        combined_handle.write(f"\n# === File: {src.name} ===\n")
                        combined_handle.write(content)
                        combined_handle.flush()
                    else:
                        self._write_output(out, content)
                    self.stats['processed'] += 1
                except Exception as e:
                    print(f"  错误: {e}")
                    self.failures.append((src, str(e)))
                    self.stats['failed'] += 1
                    continue
        finally:
            if combined_handle is not None:
                combined_handle.close()

        return self._report()

    def _report(self) -> Tuple[bool, str]:
        duration = (datetime.now() - self.stats['start_time']).total_seconds()
        print()
        print("=== 完成 ===")
        print(f"总文件:   {self.stats['total']}")
        print(f"成功:     {self.stats['processed']}")
        print(f"跳过:     {self.stats['skipped']}")
        print(f"失败:     {self.stats['failed']}")
        print(f"耗时:     {duration:.2f}s")
        if self.combine:
            print(f"输出:     {self.output}")
        if self.failures:
            print("\n失败列表:")
            for f, err in self.failures:
                print(f"  {f}: {err}")

        if self.stats['total'] == 0:
            return False, "未找到可处理的文件"
        if self.stats['processed'] > 0:
            return True, f"成功处理 {self.stats['processed']}/{self.stats['total']} 个文件"
        if self.stats['failed'] > 0:
            return False, f"所有文件处理失败 ({self.stats['failed']} 个)"
        return True, f"所有 {self.stats['total']} 个文件均已存在，未处理新文件"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description='Extract text from PDF/XLS/XLSX files to .txt files (originals untouched).',
    )
    parser.add_argument('source', nargs='*', help='Source file or folder(s) to process')
    parser.add_argument('-o', '--output', help='Output directory (per-file mode) or output file (--combine mode)')
    parser.add_argument('--no-subfolders', action='store_true', help='Do not recurse into subfolders')
    parser.add_argument('--combine', action='store_true', help='Concatenate all extracted text into a single .txt file')
    parser.add_argument('--overwrite', action='store_true', help='Overwrite existing output files')
    parser.add_argument('--list', action='store_true', help='List discoverable files without extracting')
    parser.add_argument('--ocr', action='store_true', help='Use OCR for image-only PDFs (default engine: tesseract; --engine paddleocr for Chinese small-text recall)')
    parser.add_argument('--engine', default='tesseract', choices=['tesseract', 'paddleocr'],
                        help='OCR engine when --ocr is set. Tesseract (default, ~700MB system deps) is good for typed documents. '
                             'PaddleOCR (~250MB pip + 100MB model, better on Chinese small-text) needs `pip install -r requirements-paddleocr.txt`.')
    parser.add_argument('--use-gpu', action='store_true',
                        help='Enable GPU for PaddleOCR (Win/Linux + CUDA only; ignored on macOS Apple Silicon).')
    parser.add_argument('--lang', default='chi_sim+eng', help='OCR language packs (Tesseract format; PaddleOCR maps chi_sim->ch, eng->en)')
    parser.add_argument('--dpi', type=int, default=300, help='OCR render DPI, 150-400 (default: 300; use 300+ for small text in technical drawings)')
    parser.add_argument('--rotation', type=int, default=None, choices=[0, 90, 180, 270],
                        help='Force PDF page rotation in degrees (CW=positive). Default: auto-detect by comparing OCR of original vs rotated. Ignored by PaddleOCR (rotates internally).')
    args = parser.parse_args()
    if not (150 <= args.dpi <= 400):
        parser.error(f"--dpi must be 150-400, got {args.dpi}")
    return args


def main():
    args = parse_args()

    if not args.source:
        try:
            raw = input("请输入源文件或文件夹路径 (多个用空格分隔，可使用引号): ").strip()
        except EOFError:
            raw = ""
        args.source = shlex.split(raw) if raw else []

    if not args.source:
        print("错误: 必须提供源文件或文件夹路径")
        return

    extractor = TextExtractor(
        sources=args.source,
        output=args.output,
        include_subfolders=not args.no_subfolders,
        combine=args.combine,
        overwrite=args.overwrite,
        ocr=args.ocr,
        lang=args.lang,
        dpi=args.dpi,
        rotation=args.rotation,
        engine=args.engine,
        use_gpu=args.use_gpu,
    )

    if args.list:
        valid, msg = extractor.validate()
        if not valid:
            print(f"错误: {msg}")
            sys.exit(1)
        items = extractor.discover()
        if not items:
            print("未找到可处理的文件")
            return
        print(f"找到 {len(items)} 个可处理文件:")
        for src, out, base in items:
            base = base or src.parent
            rel = src.relative_to(base)
            print(f"  [{base}] {rel}  ->  {out}")
        return

    success, message = extractor.convert()
    if success:
        print(f"\n[OK] {message}")
    else:
        print(f"\n[FAIL] {message}")
        sys.exit(1)


if __name__ == '__main__':
    main()
