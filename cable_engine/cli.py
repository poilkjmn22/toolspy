"""cable_engine.cli — V6 topology-based pipeline.

  Loader            — DWG (PDF support deferred)
  TopologyStage     — builds cable_topology table (per-analyzer dispatch)

cable→conductor→terminal topology is pre-built at scan time and
stored in cable_topology. The viewer does a direct SQL lookup.

Usage:
  python -m cable_engine.cli scan --input <dir> [--db <cable.db>]
  python -m cable_engine.cli check-cable-cabinets --db <cable.db> [--output-dir <out>]
  python -m cable_engine.cli group-by-cables --group-by-cables <xlsx> \\
      --db <cable.db> [--output-dir <out>]
"""

from __future__ import annotations

import argparse
import multiprocessing as mp
import sys
import time
from pathlib import Path
from typing import Any, Optional

from cable_engine.graph import TopologyStage
from cable_engine.ir import DocumentType
from cable_engine.loaders import get_loader_for
from cable_engine.loaders.dwg_loader import _content_hash
from cable_engine.pipeline import Context, Pipeline
from cable_engine.storage import CableStore, open_db, ensure_schema


DEFAULT_DB_FILENAME = 'cable.db'


def _discover_documents(input_dir: Path):
    """Yield (file_path, loader) for every supported DWG/DXF document
    under `input_dir`. PDFs are NOT supported in V5 P0."""
    skip_suffixes = {'.db', '.db-shm', '.db-wal', '.json', '.csv',
                     '.un~', '.swp', '.log', '.err',
                     '.pdf'}  # PDF support deferred
    # Skip names that look like sqlite journal files
    skip_names = {'cable.db-shm', 'cable.db-wal'}
    for p in sorted(input_dir.rglob('*')):
        if not p.is_file():
            continue
        if p.name.startswith('.'):
            continue
        if p.name in skip_names:
            continue
        if p.suffix.lower() in skip_suffixes:
            continue
        loader = get_loader_for(p)
        if loader is None:
            continue
        yield p, loader


def _pipeline_for(store: CableStore) -> Pipeline:
    """V6 pipeline: Loader (already done) -> TopologyStage.
    TopologyStage dispatches to the appropriate analyzer per doc type.
    """
    return Pipeline([
        TopologyStage(store),
    ])


# ---------------------------------------------------------------------------
# Per-document worker
# ---------------------------------------------------------------------------
def _process_one_document(
    document_path_str: str,
    db_path_str: str,
) -> dict:
    """Process one document in a worker process."""
    document_path = Path(document_path_str)
    loader = get_loader_for(document_path)
    if loader is None:
        return {
            'path': str(document_path),
            'error': f'no loader for {document_path.suffix}',
        }
    doc = loader.load(document_path)
    if doc is None or not doc.entities:
        return {
            'path': str(document_path),
            'content_hash': '',
            'source_type': 'unknown',
            'pages': 0,
            'entities': 0,
            'error': 'no entities loaded',
        }

    store = CableStore(open_db(Path(db_path_str), read_only=False))
    try:
        # V6.5: do the initial insert without classification; the
        # TopologyStage will set it. We update again after the pipeline
        # finishes (see below).
        store.upsert_document(
            doc.content_hash, str(document_path),
            file_size=document_path.stat().st_size if document_path.exists() else None,
            file_mtime=document_path.stat().st_mtime if document_path.exists() else None,
            document_type=doc.document_type.value,
        )
        ctx = Context(
            document_path=document_path,
            content_hash=doc.content_hash,
            document=doc,
        )
        out = _pipeline_for(store).run(ctx)
        # V6.5: persist classification back to documents row.
        if ctx.classification is not None:
            store.upsert_document(
                doc.content_hash, str(document_path),
                file_size=document_path.stat().st_size if document_path.exists() else None,
                file_mtime=document_path.stat().st_mtime if document_path.exists() else None,
                document_type=doc.document_type.value,
                classification_primary=ctx.classification.primary.value,
                classification_confidence=ctx.classification.confidence,
            )
        return {
            'path': str(document_path),
            'content_hash': doc.content_hash,
            'source_type': doc.document_type.value,
            'pages': len(doc.pages),
            'entities': len(doc.entities),
            'classification': (
                out.result.get('classification_primary', '')
                if out.result else ''
            ),
            'classification_confidence': (
                out.result.get('classification_confidence', 0.0)
                if out.result else 0.0
            ),
            'error': out.error_msg,
        }
    finally:
        store.close()


def _process_one_document_wrapper(item):
    return _process_one_document(*item)


# ---------------------------------------------------------------------------
# Main: scan subcommand
# ---------------------------------------------------------------------------
def cmd_scan(args: argparse.Namespace) -> int:
    input_dir = Path(args.input).expanduser()
    if not input_dir.is_dir():
        print(f'ERROR: input directory not found: {input_dir}', file=sys.stderr)
        return 1

    db_path = Path(args.db).expanduser() if args.db else input_dir / DEFAULT_DB_FILENAME
    db_path.parent.mkdir(parents=True, exist_ok=True)
    ensure_schema(open_db(db_path))

    print(f'Input:  {input_dir}', flush=True)
    print(f'DB:     {db_path}', flush=True)

    store = CableStore(open_db(db_path, read_only=False))
    store.set_state('started_at', time.strftime('%Y-%m-%dT%H:%M:%S'))
    store.set_state('input', str(input_dir))
    store.close()

    documents = list(_discover_documents(input_dir))
    if not documents:
        print(f'No supported DWG/DXF documents found under {input_dir}', flush=True)
        return 0
    print(f'Discovered: {len(documents)} supported documents', flush=True)

    # Resume: skip documents already in the DB with a valid classification.
    store = CableStore(open_db(db_path, read_only=True))
    done_hashes = set(
        r['content_hash'] for r in store._conn.execute(
            "SELECT content_hash FROM documents WHERE classification_primary IS NOT NULL"
        ).fetchall()
    )
    store.close()
    skipped = 0
    resume_docs: list = []
    for p, ld in documents:
        h = _content_hash(p)
        if h in done_hashes:
            skipped += 1
        else:
            resume_docs.append((p, ld))
    if skipped:
        print(f'Resume: {skipped} documents already scanned, {len(resume_docs)} remaining', flush=True)
    elif done_hashes:
        # All in DB but none have classification — likely a prior interrupted scan.
        print(f'Note: {len(done_hashes)} docs in DB but unclassified, will re-process', flush=True)

    completed = 0
    start = time.time()
    print(f'Processing with {args.workers} workers...', flush=True)

    scan_total = len(resume_docs)
    if args.workers <= 1:
        # Single-process — easier to debug
        for p, _ in resume_docs:
            res = _process_one_document(str(p), str(db_path))
            completed += 1
            elapsed = time.time() - start
            if res.get('error'):
                print(f'[{completed}/{scan_total}] {p.name} ({elapsed:.1f}s)  '
                      f'错误: {res["error"]}', flush=True)
            else:
                cls = res.get('classification', '')
                cls_short = cls.replace('_', '')[:8] if cls else '-'
                conf = res.get('classification_confidence', 0.0)
                print(f'[{completed}/{scan_total}] {p.name} ({elapsed:.1f}s)  '
                      f'[{res["source_type"]}] {res["pages"]}p/{res["entities"]}e '
                      f'cls={cls_short}({conf:.2f})',
                      flush=True)
    else:
        with mp.Pool(processes=args.workers) as pool:
            for result in pool.imap_unordered(
                _process_one_document_wrapper,
                [(str(p), str(db_path)) for p, _ in resume_docs],
            ):
                completed += 1
                elapsed = time.time() - start
                if result.get('error'):
                    print(f'[{completed}/{scan_total}] '
                          f'{Path(result["path"]).name} ({elapsed:.1f}s)  '
                          f'错误: {result["error"]}', flush=True)
                else:
                    cls = result.get('classification', '')
                    cls_short = cls.replace('_', '')[:8] if cls else '-'
                    conf = result.get('classification_confidence', 0.0)
                    print(f'[{completed}/{scan_total}] '
                          f'{Path(result["path"]).name} ({elapsed:.1f}s)  '
                          f'[{result["source_type"]}] {result["pages"]}p/{result["entities"]}e '
                          f'cls={cls_short}({conf:.2f})',
                          flush=True)

    # Summary
    store = CableStore(open_db(db_path, read_only=True))
    s = store.stats()
    store.close()
    duration = time.time() - start
    print()
    print('=== 完成 ===')
    print(f'扫描: {s["documents"]} documents')
    print(f'  topology: {s["cable_topology"]} cable-conductor records')
    print(f'  strips: {s["terminal_strips"]} terminal strips')
    print()
    # V6.5: classification breakdown
    by_cls = s.get('documents_by_classification', {})
    if by_cls:
        print('  按业务分类:')
        for cls, n in by_cls.items():
            label = {
                'circuit_loop': '回路图',
                'terminal_strip': '端子排图',
                'cable_schedule': '电缆清册',
                'protection_diagram': '保护原理图',
                'panel_layout': '屏面布置图',
                'monitoring_system': '状态监测/通风',
                'manufacturer_catalog': '厂家图册',
                'unknown': '目录/封面',
                'unclassified': '(未分类)',
            }.get(cls, cls)
            print(f'    {label:<14} {n:>4}')
        unmatched = s.get('unmatched_documents', 0)
        if unmatched:
            print(f'  无对应 analyzer 的图档: {unmatched}')
    print(f'耗时: {duration:.1f}s')
    print(f'DB: {db_path}')
    print()
    print('Start the viewer with:')
    print(f'  python -m tools.cable_match_viewer.server --db {db_path}')
    return 0


# ---------------------------------------------------------------------------
# check-cable-cabinets
# ---------------------------------------------------------------------------
def cmd_check_cable_cabinets(args: argparse.Namespace) -> int:
    """Export cable-cabinets.xlsx from cable.db.

    Generates an XLSX with columns:
      电缆编号, 电缆型号及截面,
      cable_core_numbers, cable_core_cross_section,
      cable_core_numbers2, cable_core_cross_section2,
      起点, origin_cubicle, origin_device,
      终点, terminal_cubicle, terminal_device,
      序号

    Each row corresponds to one cable (grouped by cable_id).
    电缆型号及截面: for circuit_loop documents, found by spatial
    lookup in text_entities — two TEXT blocks directly below the
    cable_id label are concatenated.  Other doc types get empty.
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        print('ERROR: openpyxl is required. pip install openpyxl', file=sys.stderr)
        return 1

    db_path = Path(args.db).expanduser()
    if not db_path.is_file():
        print(f'ERROR: DB not found: {db_path}', file=sys.stderr)
        return 1

    if args.output_dir:
        output_dir = Path(args.output_dir).expanduser().resolve()
    else:
        output_dir = db_path.parent
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / 'cable-cabinets.xlsx'

    conn = open_db(db_path, read_only=True)

    # ── Phase 1: gather per-cable data ──
    all_rows = conn.execute("""
        SELECT cable_id, cabinet_name, cabinet_name_remote,
               document_hash, source_type
        FROM cable_topology
        ORDER BY cable_id, conductor_no
    """).fetchall()

    if not all_rows:
        print('No data found in cable_topology table', file=sys.stderr)
        conn.close()
        return 1

    cables: dict[str, dict] = {}
    cable_conductor_counts: dict[str, int] = {}
    for r in all_rows:
        cid = r['cable_id']
        cable_conductor_counts[cid] = cable_conductor_counts.get(cid, 0) + 1
        if cid in cables:
            continue
        cables[cid] = {
            'cabinet_name': r['cabinet_name'] or '',
            'cabinet_name_remote': r['cabinet_name_remote'] or '',
            'cl_doc_hash': r['document_hash'] if r['source_type'] == 'circuit_loop' else None,
        }

    # ── Phase 2: load text_entities for circuit_loop documents only ──
    import re
    _CL_DOCS = {info['cl_doc_hash'] for info in cables.values()
                if info['cl_doc_hash']}
    _XS_RE = re.compile(r'(\d+)[×x*](\d+(?:\.\d+)?)')

    doc_texts: dict[str, list[dict]] = {}
    for dh in _CL_DOCS:
        doc_texts[dh] = [
            dict(r) for r in conn.execute(
                "SELECT text, x, y, entity_type "
                "FROM text_entities WHERE document_hash = ?",
                (dh,),
            ).fetchall()
        ]
    conn.close()

    # ── Phase 3: for each cable, grab the two TEXT below cable_id ──
    def _find_two_below(
        texts: list[dict], cable_id: str,
    ) -> tuple[str, str]:
        """Return (concatenated_text, cross_section).

        Find the cable_id text anchor, then collect TEXT entities
        directly below it (same x-range, lower y).  Concatenate the
        first two non-empty texts encountered when scanning downward.
        """
        anchor = None
        for t in texts:
            txt = t['text'].strip()
            if txt == cable_id:
                anchor = t
                break
        if anchor is None:
            for t in texts:
                txt = t['text'].strip()
                # Strip "(C1)" / "(1)" suffixes common in cable IDs
                paren = txt.find('(')
                bare = txt[:paren].strip() if paren > 0 else txt
                if bare == cable_id:
                    anchor = t
                    break
        if anchor is None:
            return '', ''

        ax, ay = anchor['x'], anchor['y']

        below = [
            t for t in texts
            if t is not anchor
            and abs(t['x'] - ax) < 300
            and t['y'] < ay
        ]
        below.sort(key=lambda t: t['y'], reverse=True)  # closest to anchor first

        parts: list[str] = []
        xs = ''
        for t in below:
            txt = t['text'].strip()
            if not txt or txt == cable_id:
                continue
            parts.append(txt)
            if not xs:
                m = _XS_RE.search(txt)
                if m:
                    xs = m.group(2)
            if len(parts) == 2:
                break

        return ' '.join(parts), xs

    # ── Phase 4: write XLSX ──
    wb = Workbook()
    ws = wb.active
    ws.title = 'Cable Cabinets'

    headers = [
        '电缆编号', '电缆型号及截面',
        'cable_core_numbers', 'cable_core_cross_section',
        'cable_core_numbers2', 'cable_core_cross_section2',
        '起点', 'origin_cubicle', 'origin_device',
        '终点', 'terminal_cubicle', 'terminal_device',
        '序号',
    ]
    ws.append(headers)

    for seq, (cid, info) in enumerate(sorted(cables.items()), 1):
        model_text = ''
        xs = ''
        dh = info['cl_doc_hash']
        if dh:
            texts = doc_texts.get(dh, [])
            model_text, xs = _find_two_below(texts, cid)

        ws.append([
            cid,
            model_text,
            cable_conductor_counts[cid],
            xs,
            '',
            '',
            info['cabinet_name'],
            info['cabinet_name'],
            '',
            info['cabinet_name_remote'],
            info['cabinet_name_remote'],
            '',
            seq,
        ])

    wb.save(out_path)
    print(f'已生成: {out_path}')
    print(f'共 {len(cables)} 条电缆记录')
    return 0


# ---------------------------------------------------------------------------
# group-by-cables
# ---------------------------------------------------------------------------
def cmd_group_by_cables(args: argparse.Namespace) -> int:
    """Group DWG files by cabinet using the scanned DB + XLSX schedule.

    After `python -m cable_engine.cli scan --input <dir>` has populated
    cable.db, this command queries cable_topology (via cable_id) to find
    which document each cable belongs to, then copies those DWG files into
    folders named after cubicle_model_name.

    XLSX columns: cable_nums (电缆编号, comma-separated with optional
    (cx) core suffixes to strip), cubicle_model_name (柜体型号, folder).
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print('ERROR: openpyxl is required. pip install openpyxl', file=sys.stderr)
        return 1

    import shutil

    xlsx_path = Path(args.group_by_cables).expanduser()
    db_path = Path(args.db).expanduser()
    output_dir = Path(args.output_dir or '.').expanduser().resolve()

    if not xlsx_path.is_file():
        print(f'ERROR: XLSX not found: {xlsx_path}', file=sys.stderr)
        return 1
    if not db_path.is_file():
        print(f'ERROR: DB not found: {db_path}', file=sys.stderr)
        return 1

    output_dir.mkdir(parents=True, exist_ok=True)

    # ---- Step 1: read XLSX, collect all cable IDs ----
    wb = load_workbook(xlsx_path)
    ws = wb.active

    headers: dict[str, int] = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column

    cable_col = headers.get('cable_nums') or headers.get('电缆编号')
    cubicle_col = headers.get('cubicle_model_name') or headers.get('柜体型号')

    if not cable_col or not cubicle_col:
        print('ERROR: XLSX must have cable_nums (电缆编号) and '
              'cubicle_model_name (柜体型号) columns', file=sys.stderr)
        return 1

    # Each row: (list[cable_id], cubicle_name)
    row_data: list[tuple[list[str], str]] = []
    all_cable_ids: set[str] = set()

    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=cable_col).value
        cubicle = ws.cell(row=row, column=cubicle_col).value
        if not val or not cubicle:
            continue
        parts = [p.strip() for p in str(val).strip().replace('，', ',').split(',') if p.strip()]
        ids = [p.split('(')[0].strip() for p in parts if p.split('(')[0].strip()]
        if not ids:
            continue
        row_data.append((ids, str(cubicle).strip()))
        for cid in ids:
            all_cable_ids.add(cid)

    if not row_data:
        print('ERROR: no data rows found in XLSX', file=sys.stderr)
        return 1

    print(f'XLSX: {len(row_data)} rows, {len(all_cable_ids)} unique cable IDs',
          flush=True)

    # ---- Step 2: query text_entities for cable→document mapping ----
    from .storage.sqlite import open_db
    conn = open_db(db_path, read_only=True)

    cable_to_doc: dict[str, tuple[str, str]] = {}

    # Try text_entities first (fast path — data already scanned)
    for cid in all_cable_ids:
        rows = conn.execute(
            """SELECT DISTINCT d.rel_path, d.content_hash
               FROM text_entities te
               JOIN documents d ON d.content_hash = te.document_hash
               WHERE te.text LIKE ?""",
            (f'%{cid}%',),
        ).fetchall()
        for rel_path, doc_hash in rows:
            if cid not in cable_to_doc:
                cable_to_doc[cid] = (rel_path, doc_hash)
                break

    # If text_entities is empty or only contains error messages, fall back
    # to scanning source DWGs directly via dwgread.
    _te_check = conn.execute(
        "SELECT text FROM text_entities WHERE text NOT LIKE '<%' LIMIT 1"
    ).fetchone()
    text_entities_have_data = _te_check is not None
    conn.close()

    if len(cable_to_doc) == 0 and not text_entities_have_data:
        print('text_entities has no real data (scan likely failed). '
              'Falling back to direct dwgread scan…', flush=True)

        import subprocess
        import re

        dwgread_path = shutil.which('dwgread')
        if dwgread_path is None:
            print('ERROR: dwgread not found. Install libredwg.', file=sys.stderr)
            return 1

        conn2 = open_db(db_path, read_only=True)
        doc_files = conn2.execute(
            "SELECT DISTINCT rel_path, content_hash FROM documents"
        ).fetchall()
        conn2.close()

        _TEXT_PATTERN = re.compile(r'"text"\s*:\s*"([^"]*)"')
        for row in doc_files:
            src_path = Path(row['rel_path'])
            if not src_path.is_file():
                continue
            try:
                proc = subprocess.run(
                    [dwgread_path, '-O', 'JSON', str(src_path)],
                    capture_output=True, timeout=120,
                )
            except (subprocess.TimeoutExpired, OSError):
                continue
            if proc.returncode != 0 or not proc.stdout:
                continue
            raw = proc.stdout.decode('utf-8', errors='replace')
            found_in_file: set[str] = set()
            for m in _TEXT_PATTERN.finditer(raw):
                txt = m.group(1)
                if txt in found_in_file:
                    continue
                _bare = txt.split('(')[0].strip() if '(' in txt else txt
                if _bare in all_cable_ids:
                    cable_to_doc.setdefault(_bare, ())
                    if cable_to_doc[_bare] == ():
                        cable_to_doc[_bare] = (row['rel_path'], row['content_hash'])
                    found_in_file.add(txt)

    print(f'DB: {len(cable_to_doc)}/{len(all_cable_ids)} cable IDs found',
          flush=True)

    # ---- Step 3: resolve source-file paths and process rows ----
    src_root = db_path.parent  # DWG files are relative to DB's directory
    matched = 0

    for ids, cubicle in row_data:
        found = False
        for cid in ids:
            entry = cable_to_doc.get(cid)
            if entry is None:
                continue
            rel_path, _ = entry
            src = Path(rel_path)
            if not src.is_absolute():
                src = src_root / src
            if not src.is_file():
                print(f'  WARN: {cid} maps to {src} but file not found',
                      file=sys.stderr)
                continue
            cubicle_dir = output_dir / f'{cubicle}+{cid}'
            cubicle_dir.mkdir(parents=True, exist_ok=True)
            dst = cubicle_dir / src.name
            if not dst.exists():
                shutil.copy2(src, dst)
            print(f'  {cid:20s} -> {cubicle}+{cid}/{src.name}')
            matched += 1
            found = True
            break

        if not found:
            cable_str = ', '.join(ids)
            print(f'  {"[NO MATCH]":20s} cable={cable_str} cubicle={cubicle}')

    print(f'\nDone: {matched} rows matched, '
          f'{len(row_data) - matched} rows unmatched', flush=True)
    return 0 if matched == len(row_data) else 1


# ---------------------------------------------------------------------------
# Argparse
# ---------------------------------------------------------------------------
def _build_argparser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description='cable_engine v5: graph-based document intelligence',
    )
    sub = p.add_subparsers(dest='command', required=True)

    s = sub.add_parser('scan', help='Scan a directory tree and build graph')
    s.add_argument('--input', required=True, help='Input directory (DWG/DXF)')
    s.add_argument('--db', help='Output cable.db path (default: <input>/cable.db)')
    s.add_argument('--workers', type=int, default=1, help='Parallel workers (default: 1)')

    cc = sub.add_parser(
        'check-cable-cabinets',
        help='Export cable-cabinets.xlsx from cable.db',
    )
    cc.add_argument('--db', required=True, help='cable.db path')
    cc.add_argument('--output-dir', default=None,
                    help='Output directory (default: same dir as cable.db)')

    gc = sub.add_parser(
        'group-by-cables',
        help='Group DWG files by cabinet using an XLSX cable schedule',
    )
    gc.add_argument('--group-by-cables', required=True,
                    help='XLSX file with cable_nums and cubicle_model_name columns')
    gc.add_argument('--db', required=True,
                    help='cable.db path (from a prior scan)')
    gc.add_argument('--output-dir', default=None,
                    help='Output directory (default: current dir)')
    return p


def main(argv: Optional[list[str]] = None) -> int:
    parser = _build_argparser()
    args = parser.parse_args(argv if argv is not None else sys.argv[1:])
    if args.command == 'scan':
        return cmd_scan(args)
    if args.command == 'check-cable-cabinets':
        return cmd_check_cable_cabinets(args)
    if args.command == 'group-by-cables':
        return cmd_group_by_cables(args)
    parser.print_help()
    return 1


if __name__ == '__main__':
    sys.exit(main())
